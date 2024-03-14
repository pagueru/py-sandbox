import openpyxl

# Adiciona comntários na célula de uma planilha 'X' com base em um 'de para' de outra planilha 'Y'
def adicionar_comentarios():
    
    # Carregar o arquivo Excel
    workbook = openpyxl.load_workbook(r'C:\Users\rapha\Documents\GitHub\py-sandbox\data\test_openpyxl\arquivo_openpyxl.xlsx')
    
    # Acessar as planilhas
    planilha_raiz = workbook['PlanilhaRaiz']
    planilha_comentarios = workbook['PlanilhaComentarios']
    
    # Criar um dicionário de comentários
    comentarios_dict = {row[0].value: row[1].value for row in planilha_comentarios.iter_rows(min_row=2, max_row=planilha_comentarios.max_row, min_col=1, max_col=2)}
    
    # Iterar sobre os dados na planilha raiz
    for row in planilha_raiz.iter_rows(min_row=2, max_row=planilha_raiz.max_row, min_col=1, max_col=2):
        
        # Pegar o nome da campanha
        nome_campanha = row[0].value
        
        # Verificar se há um comentário correspondente
        if nome_campanha in comentarios_dict:
            
            comentario = comentarios_dict[nome_campanha]
                        
            # Coluna onde está o NOME_CAMPANHA
            celula = row[0]
            
            # Define autor e texto do comentário
            autor = 'Raphael Coelho'          
            texto_comentario = f'{autor}: {comentario}'
            
            # Adicionar o comentário na célula correspondente na planilha raiz
            comment = openpyxl.comments.Comment(texto_comentario, 'Raphael Coelho')

            # Adiciona o comentário à célula
            celula.comment = comment
    
    # Salvar as alterações no mesmo arquivo
    workbook.save(r'C:\Users\rapha\Documents\GitHub\py-sandbox\data\test_openpyxl\arquivo_openpyxl.xlsx')

# Chamando a função para adicionar os comentários
adicionar_comentarios()